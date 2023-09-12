import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill
import reader
from collections import OrderedDict


class DataProcessor:
    def __init__(self, filename):
        self.filename = filename

    def start_process(self):
        header_style = main_header_style()
        data = process_data(self.filename)
        #create_excel(data, header_style)


def process_data(filename):
    data = []
    # Get the vcf records
    vcf_records = reader.read_vcf(filename)

    # For each record create a dictionary and add it to the list called data.
    # if you are particular about order you can use OrderedDict otherwise a normal dict
    for record in vcf_records:
        temp_dict = OrderedDict()
        temp_dict["dbsnp rsid"] = record.ID
        temp_dict["Chromosome"] = record.CHROM
        temp_dict["Position"] = record.POS
        temp_dict["Reference Allele"] = str(record.REF)
        temp_dict["Alternate Allele"] = str(record.ALT)
        temp_dict["Quality"] = record.QUAL
        if record.FILTER and record.FILTER != []:
            temp_dict["Filter Status"] = record.FILTER
        else:
            temp_dict["Filter Status"] = "PASS"
        info_keys = [x for x in vcf_records.infos]
        i = 0
        for key, value in record.INFO.items():
            if key == info_keys[i]:
                if value:
                    temp_dict[key] = str(value)
                else:
                    temp_dict[key] = "NA"
            else:
                temp_dict[info_keys[i]] = "NA"
            i += 1

        format_parts = record.FORMAT.split(":")
        for sample in record.samples:
            for i in range(len(format_parts)):
                temp_dict[format_parts[i]] = sample.data[i]
        data.append(temp_dict)
    return data


def create_excel(data, header_style):
    # Write the Header to the sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    if data:
        header = list(data[0].keys())
        sheet.append(header)
        row = sheet[1]
        for cell in row:
            cell.style = header_style
    for row in data:
        sheet.append(list(row.values()))
    wb.save("sample_excel.xlsx")


def main_header_style():
    main_header = NamedStyle(name="main_header")
    main_header.font = Font(bold=True)
    main_header.fill = PatternFill(fgColor="ffcc66", fill_type="solid")
    main_header.border = Border(bottom=Side(border_style="thick"))
    main_header.alignment = Alignment(horizontal="center")
    return main_header


if __name__ == "__main__":
    print("This is running as a standalone")
    DataProcessor("sample.vcf").start_process()
